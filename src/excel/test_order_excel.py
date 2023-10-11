import pdb

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

file_path = 'excel_templates/L1001.xlsx'


if __name__ == '__main__':
    wb = load_workbook(file_path)
    sheet = wb.active
    image_loader = SheetImageLoader(sheet)
    cell = sheet.cell(7, 2)
    pdb.set_trace()
    cell = sheet.cell(7, 1)

    pdb.set_trace()
    cell = sheet.cell(8, 1)

    pdb.set_trace()
    cell = sheet.cell(9, 1)

    pdb.set_trace()
    image_loader.get('B8').show()
    pdb.set_trace()
    image_loader.get('D7').show()
    pdb.set_trace()
    image_loader.get('D8').show()
    pdb.set_trace()
    image_loader.get('D9').show()
    pdb.set_trace()
    image_loader.get('D10').show()
    pdb.set_trace()

    cell = sheet.cell(7, 4)
    pdb.set_trace()
